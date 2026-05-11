from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.core.models import (
    Empresa,
    GerenciaSINFRA,
    Predio,
    ServicoProcesso,
    SituacaoSIPAC,
    StatusProcesso,
    TipoAmbiente,
)

from .models import InteressadoProcesso, Orcamento, Processo


# ── Cabeçalhos e metadados das colunas ────────────────────────────────────────

COLUNAS = [
    # (nome_coluna, campo_modelo, obrigatorio, descricao_legenda)
    ("numero_processo",   "numero_processo",  True,  "Número do processo no SIPAC (ex: 23074.001234/2024-01). Obrigatório e único."),
    ("assunto",           "assunto",          False, "Descrição resumida do objeto do processo."),
    ("data_abertura",     "data_abertura",    False, "Data de abertura do processo. Formato: DD/MM/AAAA ou AAAA-MM-DD."),
    ("data_os",           "data_os",          False, "Data de emissão da Ordem de Serviço. Formato: DD/MM/AAAA ou AAAA-MM-DD."),
    ("data_conclusao",    "data_conclusao",   False, "Data de conclusão do serviço. Formato: DD/MM/AAAA ou AAAA-MM-DD."),
    ("data_arquivamento", "data_arquivamento",False, "Data de arquivamento do processo. Formato: DD/MM/AAAA ou AAAA-MM-DD."),
    ("status",            "status",           False, "Código ou nome do status do processo. Ver aba LEGENDAS para valores válidos."),
    ("situacao_sipac",    "situacao_sipac",   False, "Situação no SIPAC: ATIVO, ARQUIVADO ou APENSADO."),
    ("gerencia",          "gerencia",         False, "Nome da gerência da SINFRA responsável. Ver aba LEGENDAS para valores válidos."),
    ("servico",           "servico",          False, "Nome do serviço do processo. Ver aba LEGENDAS para valores válidos."),
    ("predio",            "predio",           False, "Nome do prédio (deve corresponder ao cadastro). Ver aba LEGENDAS para valores válidos."),
    ("tipo_ambiente",     "tipo_ambiente",    False, "Tipo de ambiente (ex: Sala de Aula, Banheiro). Ver aba LEGENDAS para valores válidos."),
    ("empresa",           "empresa",          False, "Nome da empresa executora. Se não existir, será criada automaticamente."),
    ("classificacao_az",  "classificacao_az", False, "Letra de priorização A–Z (apenas uma letra maiúscula)."),
    ("link_sipac",        "link_sipac",       False, "URL do processo no SIPAC (https://...)."),
    ("observacao",        "observacao",       False, "Observação interna (não aparece publicamente)."),
    ("acompanhamento_ct", "acompanhamento_ct",False, "Histórico de acompanhamento interno do CT."),
    ("unidade_origem",    "unidade_origem",   False, "Unidade de origem informada no SIPAC público."),
]

HEADER_NAMES = [c[0] for c in COLUNAS]
INTERESSADOS_SHEET_NAME = "INTERESSADOS_PROCESSO"
INTERESSADOS_HEADER_NAMES = ["numero_processo", "tipo", "identificador", "nome"]

# Mapeamento de variações de cabeçalho → campo canônico
# Chaves já normalizadas (minúsculas, sem acentos, underscores)
COLUMN_ALIASES: dict[str, str] = {
    # numero_processo
    "numero_processo": "numero_processo",
    "n_processo": "numero_processo",
    "num_processo": "numero_processo",
    "numero": "numero_processo",
    "processo": "numero_processo",
    "n_do_processo": "numero_processo",
    "numero_do_processo": "numero_processo",
    "n_sipac": "numero_processo",
    "num_sipac": "numero_processo",
    "no_processo": "numero_processo",
    "n": "numero_processo",

    # assunto
    "assunto": "assunto",
    "descricao": "assunto",
    "objeto": "assunto",
    "titulo": "assunto",

    # datas
    "data_abertura": "data_abertura",
    "abertura": "data_abertura",
    "dt_abertura": "data_abertura",
    "data_de_abertura": "data_abertura",
    "data_abertura_processo": "data_abertura",

    "data_os": "data_os",
    "os": "data_os",
    "ordem_servico": "data_os",
    "data_ordem_servico": "data_os",
    "dt_os": "data_os",
    "data_emissao_os": "data_os",
    "emissao_os": "data_os",

    "data_conclusao": "data_conclusao",
    "conclusao": "data_conclusao",
    "dt_conclusao": "data_conclusao",
    "data_de_conclusao": "data_conclusao",
    "data_encerramento": "data_conclusao",

    "data_arquivamento": "data_arquivamento",
    "arquivamento": "data_arquivamento",
    "dt_arquivamento": "data_arquivamento",
    "data_de_arquivamento": "data_arquivamento",

    # status
    "status": "status",
    "status_processo": "status",
    "cod_status": "status",
    "codigo_status": "status",
    "situacao_status": "status",
    "status_sipac": "status",

    # situacao_sipac
    "situacao_sipac": "situacao_sipac",
    "situacao": "situacao_sipac",
    "sit_sipac": "situacao_sipac",
    "situacao_no_sipac": "situacao_sipac",
    "situacao_processo": "situacao_sipac",

    # gerencia
    "gerencia": "gerencia",
    "gerencia_sinfra": "gerencia",
    "gerencia_responsavel": "gerencia",
    "divisao": "gerencia",
    "setor_sinfra": "gerencia",

    # servico
    "servico": "servico",
    "servico_processo": "servico",
    "tipo_servico": "servico",
    "servico_contratado": "servico",
    "natureza_servico": "servico",

    # predio
    "predio": "predio",
    "edificio": "predio",
    "bloco": "predio",
    "local": "predio",
    "localizacao": "predio",
    "nome_predio": "predio",
    "nome_edificio": "predio",
    "local_predio": "predio",
    "local_predio_": "predio",
    "predio_local": "predio",

    # tipo_ambiente
    "tipo_ambiente": "tipo_ambiente",
    "ambiente": "tipo_ambiente",
    "tipo_de_ambiente": "tipo_ambiente",
    "tipo_local": "tipo_ambiente",
    "local_servico": "tipo_ambiente",
    "local_do_servico": "tipo_ambiente",
    "ambiente_servico": "tipo_ambiente",

    # gerencia  (divisão da SINFRA = gerência responsável)
    "gerencia": "gerencia",
    "gerencia_sinfra": "gerencia",
    "gerencia_responsavel": "gerencia",
    "divisao": "gerencia",
    "setor_sinfra": "gerencia",
    "servico_divisao": "gerencia",
    "servico_divisao_": "gerencia",
    "divisao_sinfra": "gerencia",
    "divisao_da_sinfra": "gerencia",

    # servico  (tipo de serviço)
    "servico": "servico",
    "servico_processo": "servico",
    "tipo_servico": "servico",
    "tipo_de_servico": "servico",
    "servico_contratado": "servico",
    "natureza_servico": "servico",
    "natureza_do_servico": "servico",

    # empresa
    "empresa": "empresa",
    "empresa_executora": "empresa",
    "contratada": "empresa",
    "empresa_contratada": "empresa",
    "prestadora": "empresa",
    "fornecedor": "empresa",

    # classificacao_az
    "classificacao_az": "classificacao_az",
    "az": "classificacao_az",
    "prioridade_az": "classificacao_az",
    "classificacao": "classificacao_az",
    "prioridade": "classificacao_az",
    "classe": "classificacao_az",
    "categoria": "classificacao_az",
    "classificacao_a_z": "classificacao_az",

    # link_sipac
    "link_sipac": "link_sipac",
    "link": "link_sipac",
    "sipac": "link_sipac",
    "url_sipac": "link_sipac",
    "link_do_sipac": "link_sipac",
    "url": "link_sipac",
    "endereco_sipac": "link_sipac",

    # observacao
    "observacao": "observacao",
    "obs": "observacao",
    "observacoes": "observacao",
    "notas": "observacao",
    "nota": "observacao",

    # acompanhamento_ct
    "acompanhamento_ct": "acompanhamento_ct",
    "acompanhamento": "acompanhamento_ct",
    "historico": "acompanhamento_ct",
    "historico_ct": "acompanhamento_ct",
    "andamento": "acompanhamento_ct",

    # unidade_origem
    "unidade_origem": "unidade_origem",
    "unidade_de_origem": "unidade_origem",
    "origem_unidade": "unidade_origem",
    "unidade_sipac": "unidade_origem",

    # campos especiais (tratados em _fill_processo)
    "orcamento_r": "_orcamento_valor",
    "orcamento_r_": "_orcamento_valor",
    "orcamento": "_orcamento_valor",
    "valor_orcamento": "_orcamento_valor",
    "valor": "_orcamento_valor",
    "nota_de_empenho": "_nota_empenho",
    "nota_em_empenho": "_nota_empenho",
    "nota_empenho": "_nota_empenho",
    "empenho": "_nota_empenho",
    "numero_empenho": "_nota_empenho",
    "n_empenho": "_nota_empenho",
    "n_requisicoes_origem": "_requisicoes_origem",
    "requisicoes_origem": "_requisicoes_origem",
    "n_requisicoes": "_requisicoes_origem",
    "requisicoes": "_requisicoes_origem",
    "req_origem": "_requisicoes_origem",
    "origem": "_requisicoes_origem",

    # predio — variações adicionais
    "local_da_intenvencao": "predio",   # grafia com typo da planilha
    "local_da_intervencao": "predio",
    "local_intervencao": "predio",
    "local_intenvencao": "predio",
    "localizacao_do_processo": "predio",
    "localizacao_processo": "predio",

    # link_sipac — variações adicionais
    "link_para_o_processo": "link_sipac",
    "link_processo": "link_sipac",
    "link_do_processo": "link_sipac",

    # colunas calculadas / auxiliares — reconhecidas mas ignoradas na importação
    "latitude":          "_ignorado",
    "longitude":         "_ignorado",
    "tempo_de_reacao":   "_ignorado",
    "tempo_de_execucao": "_ignorado",
    "tempo_de_analise":  "_ignorado",
    "tempo_de_solucao":  "_ignorado",
    "procv":             "_ignorado",
    "formula":           "_ignorado",
    "auxiliar":          "_ignorado",
}

# Cor do cabeçalho: vinho do módulo
COR_HEADER_FG = "FFFFFF"
COR_HEADER_BG = "7A2632"
COR_OBG_BG    = "F4D0D4"   # rosa claro para colunas obrigatórias


# ── Gerador de modelo ─────────────────────────────────────────────────────────

def gerar_modelo_xlsx() -> bytes:
    """Retorna bytes de um arquivo XLSX com a aba PROCESSOS (modelo) e LEGENDAS."""
    wb = Workbook()

    # ── aba PROCESSOS ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "PROCESSOS"

    header_font  = Font(bold=True, color=COR_HEADER_FG)
    header_fill  = PatternFill("solid", fgColor=COR_HEADER_BG)
    obg_fill     = PatternFill("solid", fgColor=COR_OBG_BG)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (nome, _, obg, _desc) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font  = header_font if obg else Font(bold=True)
        cell.fill  = header_fill if obg else obg_fill if obg else PatternFill("solid", fgColor="D9D9D9")
        cell.alignment = center_align

    # Larguras de coluna
    larguras = {
        "numero_processo": 28, "assunto": 50,
        "data_abertura": 16, "data_os": 16, "data_conclusao": 16, "data_arquivamento": 18,
        "status": 20, "situacao_sipac": 18, "gerencia": 38, "servico": 32,
        "predio": 28, "tipo_ambiente": 22, "empresa": 28,
        "classificacao_az": 14, "link_sipac": 40,
        "observacao": 40, "acompanhamento_ct": 40, "unidade_origem": 42,
    }
    for col_idx, nome in enumerate(HEADER_NAMES, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(nome, 20)
    ws.row_dimensions[1].height = 28

    # 3 linhas de exemplo comentadas (guias visuais)
    ws.cell(row=2, column=1, value="23074.001234/2024-01")
    ws.cell(row=2, column=2, value="Reforma do telhado do Bloco CTB")
    ws.cell(row=2, column=3, value="10/03/2024")
    ws.cell(row=2, column=7, value="07")
    ws.cell(row=2, column=8, value="ATIVO")
    ws.cell(row=2, column=14, value="B")
    for col_idx in range(1, len(COLUNAS) + 1):
        ws.cell(row=2, column=col_idx).font = Font(italic=True, color="888888")

    # â”€â”€ aba INTERESSADOS_PROCESSO â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    wi = wb.create_sheet(INTERESSADOS_SHEET_NAME)
    for col_idx, nome in enumerate(INTERESSADOS_HEADER_NAMES, start=1):
        cell = wi.cell(row=1, column=col_idx, value=nome)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    wi.cell(row=2, column=1, value="23074.001234/2024-01")
    wi.cell(row=2, column=2, value="Unidade")
    wi.cell(row=2, column=3, value="110055")
    wi.cell(row=2, column=4, value="CENTRO DE TECNOLOGIA (CT)")
    for col_idx in range(1, len(INTERESSADOS_HEADER_NAMES) + 1):
        wi.cell(row=2, column=col_idx).font = Font(italic=True, color="888888")
    for col_idx, width in enumerate([28, 22, 18, 55], start=1):
        wi.column_dimensions[get_column_letter(col_idx)].width = width
    wi.row_dimensions[1].height = 28

    # ── aba LEGENDAS ──────────────────────────────────────────────────────────
    wl = wb.create_sheet("LEGENDAS")
    wl.column_dimensions["A"].width = 22
    wl.column_dimensions["B"].width = 60
    wl.column_dimensions["C"].width = 55

    legend_header = Font(bold=True, color=COR_HEADER_FG)
    legend_fill   = PatternFill("solid", fgColor=COR_HEADER_BG)
    for col, titulo in enumerate(["Coluna", "Descrição", "Valores válidos / exemplos"], start=1):
        cell = wl.cell(row=1, column=col, value=titulo)
        cell.font = legend_header
        cell.fill = legend_fill
        cell.alignment = Alignment(horizontal="center")

    # Colunas com listas de valores do banco
    try:
        status_vals = " | ".join(
            f"{s.codigo} — {s.nome}" for s in StatusProcesso.objects.order_by("ordem")[:13]
        )
    except Exception:
        status_vals = "Código do status (ex: 01, 07, 08…)"

    try:
        gerencia_vals = " | ".join(g.nome for g in GerenciaSINFRA.objects.order_by("nome"))
    except Exception:
        gerencia_vals = "Nome da gerência SINFRA"

    try:
        servico_vals = " | ".join(s.nome for s in ServicoProcesso.objects.order_by("nome")[:20])
    except Exception:
        servico_vals = "Nome do serviço"

    try:
        predio_vals = " | ".join(p.nome for p in Predio.objects.order_by("nome")[:15])
    except Exception:
        predio_vals = "Nome do prédio"

    try:
        tipo_vals = " | ".join(t.nome for t in TipoAmbiente.objects.order_by("nome")[:15])
    except Exception:
        tipo_vals = "Tipo de ambiente"

    extras = {
        "status": status_vals,
        "situacao_sipac": "ATIVO | ARQUIVADO | APENSADO",
        "gerencia": gerencia_vals,
        "servico": servico_vals,
        "predio": predio_vals,
        "tipo_ambiente": tipo_vals,
        "data_abertura": "DD/MM/AAAA ou AAAA-MM-DD",
        "data_os": "DD/MM/AAAA ou AAAA-MM-DD",
        "data_conclusao": "DD/MM/AAAA ou AAAA-MM-DD",
        "data_arquivamento": "DD/MM/AAAA ou AAAA-MM-DD",
        "classificacao_az": "Uma letra maiúscula: A, B, C, D…",
        "link_sipac": "URL completa: https://sipac.ufpb.br/...",
        "numero_processo": "Formato SIPAC: 23074.XXXXXX/AAAA-DV",
    }

    for row_idx, (nome, _campo, obg, desc) in enumerate(COLUNAS, start=2):
        obg_txt = " [OBRIGATÓRIO]" if obg else ""
        wl.cell(row=row_idx, column=1, value=nome + obg_txt)
        wl.cell(row=row_idx, column=2, value=desc)
        wl.cell(row=row_idx, column=3, value=extras.get(nome, "Texto livre"))
        if obg:
            for col in range(1, 4):
                wl.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor=COR_OBG_BG)
        wl.row_dimensions[row_idx].height = 22
        wl.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
        wl.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True)

    # Salva em memória
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Importador ────────────────────────────────────────────────────────────────

class ProcessoImporter:
    """Importa processos de um arquivo XLSX/XLSM/CSV."""

    DATE_FORMATS = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"]

    def __init__(self):
        # Caches para evitar N+1 queries
        self._status_cache:    dict[str, StatusProcesso]   = {}
        self._gerencia_cache:  dict[str, GerenciaSINFRA]   = {}
        self._servico_cache:   dict[str, ServicoProcesso]  = {}
        self._situacao_cache:  dict[str, SituacaoSIPAC]    = {}
        self._predio_cache:    dict[str, Predio]            = {}
        self._ambiente_cache:  dict[str, TipoAmbiente]     = {}
        self._empresa_cache:   dict[str, Empresa]           = {}

    # ── Ponto de entrada ─────────────────────────────────────────────────────

    def import_file(self, uploaded_file) -> dict[str, Any]:
        """Importa arquivo e retorna dict com resumo."""
        suffix = Path(uploaded_file.name).suffix.lower()
        if suffix not in {".xlsm", ".xlsx"}:
            raise ValueError("Formato inválido. Envie um arquivo XLSX ou XLSM.")

        uploaded_file.seek(0)
        wb = load_workbook(
            uploaded_file,
            read_only=True,
            data_only=True,
            keep_vba=(suffix == ".xlsm"),
        )

        # Aceita aba "PROCESSOS" ou a primeira aba
        sheet = None
        for name in wb.sheetnames:
            if name.strip().upper() == "PROCESSOS":
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active

        rows, col_info = self._sheet_to_dicts(sheet)

        interessados_rows: list[dict] = []
        for name in wb.sheetnames:
            if name.strip().upper() == INTERESSADOS_SHEET_NAME:
                interessados_rows = self._interessados_sheet_to_dicts(wb[name])
                break

        with transaction.atomic():
            resultado = self._upsert_processos(rows)
            interessados_result = self._upsert_interessados(interessados_rows)
        resultado["colunas"] = col_info
        resultado.update(interessados_result)
        return resultado

    # ── Conversão de planilha ─────────────────────────────────────────────────

    def _sheet_to_dicts(self, sheet) -> tuple[list[dict], list[dict]]:
        """Retorna (linhas_como_dicts, info_colunas).
        info_colunas: lista de {original, normalizado, canonico, mapeado}."""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], []

        # Mapeia cada cabeçalho original → chave canônica
        col_info = []
        headers_canonical = []
        for h in rows[0]:
            original = str(h).strip() if h is not None else ""
            norm = self._norm(original) if original else ""
            canonical = COLUMN_ALIASES.get(norm, norm)
            ignorado = canonical == "_ignorado"
            mapeado = canonical in HEADER_NAMES or canonical.startswith("_")
            col_info.append({
                "original": original,
                "normalizado": norm,
                "canonico": canonical,
                "mapeado": mapeado,    # True p/ campos mapeados E ignorados
                "ignorado": ignorado,  # True apenas p/ colunas calculadas/auxiliares
            })
            headers_canonical.append(canonical)

        result = []
        for row in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            d = {}
            for i, v in enumerate(row):
                if i < len(headers_canonical):
                    d[headers_canonical[i]] = str(v).strip() if v is not None else ""
            result.append(d)

        return result, col_info

    def _interessados_sheet_to_dicts(self, sheet) -> list[dict]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = []
        for h in rows[0]:
            original = str(h).strip() if h is not None else ""
            norm = self._norm(original) if original else ""
            headers.append(COLUMN_ALIASES.get(norm, norm))

        result = []
        for row in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            d = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    d[headers[i]] = str(value).strip() if value is not None else ""
            result.append(d)
        return result

    @staticmethod
    def _norm(text: str) -> str:
        """Normaliza cabeçalho: minúsculas, sem acentos, underscores."""
        import unicodedata
        text = unicodedata.normalize("NFD", text.lower())
        text = "".join(c for c in text if unicodedata.category(c) != "Mn")
        text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
        return text

    # ── Upsert ────────────────────────────────────────────────────────────────

    def _upsert_processos(self, rows: list[dict]) -> dict[str, Any]:
        created = updated = skipped = 0
        erros: list[str] = []

        for i, row in enumerate(rows, start=2):
            num = self._get(row, "numero_processo", "numero", "n_processo", "num_processo")
            if not num:
                skipped += 1
                continue
            num = num.strip()
            try:
                processo, is_new = Processo.objects.get_or_create(numero_processo=num)
                avisos = self._fill_processo(processo, row)
                processo.save()
                if is_new:
                    created += 1
                else:
                    updated += 1
                for av in avisos:
                    erros.append(f"Linha {i} ({num}): {av}")
            except Exception as exc:
                erros.append(f"Linha {i} ({num}): {exc}")

        return {
            "criados": created,
            "atualizados": updated,
            "ignorados": skipped,
            "erros": erros,
            "total_linhas": len(rows),
        }

    def _upsert_interessados(self, rows: list[dict]) -> dict[str, Any]:
        if not rows:
            return {
                "interessados_importados": 0,
                "interessados_ignorados": 0,
                "interessados_erros": [],
            }

        imported = skipped = 0
        erros: list[str] = []
        por_processo: dict[str, list[dict]] = {}

        for i, row in enumerate(rows, start=2):
            numero = self._get(row, "numero_processo", "processo")
            nome = self._get(row, "nome")
            if not numero or not nome:
                skipped += 1
                continue
            por_processo.setdefault(numero.strip(), []).append({
                "linha": i,
                "tipo": self._get(row, "tipo") or "",
                "identificador": self._get(row, "identificador") or "",
                "nome": nome,
            })

        for numero, interessados in por_processo.items():
            processo = Processo.objects.filter(numero_processo=numero).first()
            if not processo:
                skipped += len(interessados)
                erros.append(f"Aba {INTERESSADOS_SHEET_NAME} ({numero}): processo não encontrado")
                continue

            processo.interessados.all().delete()
            seen: set[tuple[str, str, str]] = set()
            novos = []
            for item in interessados:
                key = (item["tipo"], item["identificador"], item["nome"])
                if key in seen:
                    continue
                seen.add(key)
                novos.append(InteressadoProcesso(
                    processo=processo,
                    tipo=item["tipo"],
                    identificador=item["identificador"],
                    nome=item["nome"],
                ))
            InteressadoProcesso.objects.bulk_create(novos)
            imported += len(novos)

        return {
            "interessados_importados": imported,
            "interessados_ignorados": skipped,
            "interessados_erros": erros,
        }

    def _fill_processo(self, processo: Processo, row: dict) -> list[str]:
        """Preenche campos do processo a partir da linha da planilha.
        Retorna lista de avisos (valores não encontrados etc.)."""
        avisos: list[str] = []

        def set_if(attr, *keys):
            val = self._get(row, *keys)
            if val:
                setattr(processo, attr, val)

        def set_fk(attr, resolver, *keys):
            """Só toca no campo FK se a célula tiver valor."""
            val = self._get(row, *keys)
            if val is None:
                return  # célula vazia → não sobrescreve
            obj = resolver(val)
            if obj is not None:
                setattr(processo, attr, obj)
            else:
                avisos.append(f"{keys[0]}='{val}' não encontrado")
                setattr(processo, attr, None)  # valor fornecido mas não resolveu → limpa

        # Texto livre
        set_if("assunto",           "assunto")
        set_if("link_sipac",        "link_sipac", "link", "sipac")
        set_if("observacao",        "observacao", "obs")
        set_if("acompanhamento_ct", "acompanhamento_ct", "acompanhamento")
        set_if("unidade_origem",    "unidade_origem", "unidade_sipac")

        # Classificação A-Z
        az = self._get(row, "classificacao_az", "az", "prioridade_az", "classificacao")
        if az:
            az = az.upper().strip()
            if len(az) == 1 and az.isalpha():
                processo.classificacao_az = az

        # Datas
        for attr, *keys in [
            ("data_abertura",     "data_abertura",     "abertura"),
            ("data_os",           "data_os",            "data_ordem_servico", "os"),
            ("data_conclusao",    "data_conclusao",     "conclusao"),
            ("data_arquivamento", "data_arquivamento",  "arquivamento"),
        ]:
            val = self._get(row, attr, *keys)
            if val:
                parsed = self._parse_date(val)
                if parsed:
                    setattr(processo, attr, parsed)
                else:
                    avisos.append(f"{attr}='{val}' não reconhecido como data")

        # FK por nome/código — só altera se célula tiver valor
        set_fk("status",         self._resolve_status,   "status",        "status_processo")
        set_fk("gerencia",       self._resolve_gerencia, "gerencia",      "gerencia_sinfra")
        set_fk("servico",        self._resolve_servico,  "servico",       "servico_processo", "tipo_servico")
        set_fk("situacao_sipac", self._resolve_situacao, "situacao_sipac","situacao", "sit_sipac")
        set_fk("predio",         self._resolve_predio,   "predio",        "edificio", "bloco")
        set_fk("tipo_ambiente",  self._resolve_ambiente, "tipo_ambiente", "ambiente")
        set_fk("empresa",        self._resolve_empresa,  "empresa",       "empresa_executora")

        # ── Campos especiais ──────────────────────────────────────────────────

        # Orçamento (valor R$)
        valor_orc_raw = self._get(row, "_orcamento_valor")
        if valor_orc_raw:
            valor_dec = self._parse_decimal(valor_orc_raw)
            if valor_dec is not None:
                # Atualiza orçamento marcado como "importado" ou cria novo
                orc = processo.orcamentos.filter(descricao="importado").first()
                if orc:
                    orc.valor = valor_dec
                    orc.save(update_fields=["valor"])
                else:
                    next_seq = (processo.orcamentos.count() or 0) + 1
                    Orcamento.objects.create(
                        processo=processo,
                        numero_sequencial=next_seq,
                        descricao="importado",
                        valor=valor_dec,
                    )
            else:
                avisos.append(f"orcamento='{valor_orc_raw}' não reconhecido como valor numérico")

        # Nota de empenho → preserva no campo observacao se ainda não registrado
        nota_emp_raw = self._get(row, "_nota_empenho")
        if nota_emp_raw:
            tag = f"[Empenho: {nota_emp_raw}]"
            if tag not in (processo.observacao or ""):
                processo.observacao = (
                    (processo.observacao + "\n" + tag).strip()
                    if processo.observacao
                    else tag
                )

        # Requisições de origem → vínculo M2M (processo.pk já existe: get_or_create)
        req_str = self._get(row, "_requisicoes_origem")
        if req_str:
            from apps.tracker.models import Requisicao
            codigos = [c.strip() for c in re.split(r"[;,\s]+", req_str) if c.strip()]
            for cod in codigos:
                req = Requisicao.objects.filter(codigo__iexact=cod).first()
                if req:
                    processo.requisicoes.add(req)
                else:
                    avisos.append(f"requisição origem '{cod}' não encontrada")

        return avisos

    # ── Resolvers de FK ───────────────────────────────────────────────────────

    def _resolve_status(self, val: str | None) -> StatusProcesso | None:
        if not val:
            return None
        key = val.strip()
        if key not in self._status_cache:
            # Código numérico pode vir sem zero à esquerda (ex: "7" → "07")
            padded = key.zfill(2) if key.isdigit() else None
            obj = (
                StatusProcesso.objects.filter(codigo__iexact=key).first()
                or (StatusProcesso.objects.filter(codigo__iexact=padded).first() if padded else None)
                or StatusProcesso.objects.filter(codigo__istartswith=key).first()
                or StatusProcesso.objects.filter(nome__iexact=key).first()
                or StatusProcesso.objects.filter(nome__icontains=key).first()
            )
            self._status_cache[key] = obj
        return self._status_cache[key]

    def _resolve_gerencia(self, val: str | None) -> GerenciaSINFRA | None:
        if not val:
            return None
        key = val.strip()
        if key not in self._gerencia_cache:
            chave = self._chave(key)
            self._gerencia_cache[key] = (
                GerenciaSINFRA.objects.filter(nome__iexact=key).first()
                or GerenciaSINFRA.objects.filter(nome__icontains=key).first()
                or GerenciaSINFRA.objects.filter(nome__icontains=chave).first()
            )
        return self._gerencia_cache[key]

    def _resolve_servico(self, val: str | None) -> ServicoProcesso | None:
        if not val:
            return None
        key = val.strip()
        if key not in self._servico_cache:
            chave = self._chave(key)
            self._servico_cache[key] = (
                ServicoProcesso.objects.filter(nome__iexact=key).first()
                or ServicoProcesso.objects.filter(nome__icontains=key).first()
                or ServicoProcesso.objects.filter(nome__icontains=chave).first()
            )
        return self._servico_cache[key]

    def _resolve_situacao(self, val: str | None) -> SituacaoSIPAC | None:
        if not val:
            return None
        key = val.strip().upper()
        if key not in self._situacao_cache:
            self._situacao_cache[key] = (
                SituacaoSIPAC.objects.filter(nome__iexact=key).first()
                or SituacaoSIPAC.objects.filter(nome__icontains=key).first()
            )
        return self._situacao_cache[key]

    def _resolve_predio(self, val: str | None) -> Predio | None:
        if not val:
            return None
        key = val.strip()
        if key not in self._predio_cache:
            # Tenta nome exato, depois chave_normalizada (sem acentos/espaços),
            # depois nome parcial
            chave = self._chave(key)
            self._predio_cache[key] = (
                Predio.objects.filter(nome__iexact=key).first()
                or Predio.objects.filter(chave_normalizada__iexact=chave).first()
                or Predio.objects.filter(nome__icontains=key).first()
                or Predio.objects.filter(chave_normalizada__icontains=chave).first()
            )
        return self._predio_cache[key]

    def _resolve_ambiente(self, val: str | None) -> TipoAmbiente | None:
        if not val:
            return None
        key = val.strip()
        if key not in self._ambiente_cache:
            chave = self._chave(key)
            self._ambiente_cache[key] = (
                TipoAmbiente.objects.filter(nome__iexact=key).first()
                or TipoAmbiente.objects.filter(nome__icontains=key).first()
                or TipoAmbiente.objects.filter(nome__icontains=chave).first()
            )
        return self._ambiente_cache[key]

    def _resolve_empresa(self, val: str | None) -> Empresa | None:
        if not val:
            return None
        key = val.strip()
        if key not in self._empresa_cache:
            # Busca exata primeiro, depois parcial; cria se não existir
            obj = (
                Empresa.objects.filter(nome__iexact=key).first()
                or Empresa.objects.filter(nome__icontains=key).first()
            )
            if obj is None:
                obj = Empresa.objects.create(nome=key, ativa=True)
            self._empresa_cache[key] = obj
        return self._empresa_cache[key]

    # ── Utilitários ───────────────────────────────────────────────────────────

    @staticmethod
    def _get(row: dict, *keys: str) -> str | None:
        for key in keys:
            # Tenta exato e normalizado
            val = row.get(key) or row.get(ProcessoImporter._norm(key))
            if val and str(val).strip():
                return str(val).strip()
        return None

    def _parse_date(self, val) -> date | None:
        if not val:
            return None
        # openpyxl pode retornar datetime/date direto (antes da conversão para string)
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        val_str = str(val).strip()
        if not val_str or val_str.lower() in ("none", "nan", "nat"):
            return None
        # openpyxl serializado como string traz "AAAA-MM-DD HH:MM:SS" — descarta horário
        if " " in val_str:
            val_str = val_str.split(" ")[0]
        # Remove frações de segundo se presentes (ex: "2024-03-10T00:00:00")
        val_str = val_str.replace("T", " ").split(" ")[0]
        for fmt in self.DATE_FORMATS:
            try:
                return datetime.strptime(val_str, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _chave(text: str) -> str:
        """Reproduz normalize_key() do domain.py: NFKD, sem acentos, sem espaços."""
        import unicodedata as _ud
        t = _ud.normalize("NFKD", text)
        t = "".join(ch for ch in t if not _ud.combining(ch))
        return re.sub(r"\s+", "", t)

    @staticmethod
    def _parse_decimal(val: str) -> Decimal | None:
        """Converte string monetária em Decimal. Aceita formatos BR e US."""
        s = str(val).strip()
        # Remove símbolos de moeda, espaços e aspas
        s = re.sub(r"[R$\s\"']", "", s)
        if not s or s.lower() in ("none", "nan", "-"):
            return None
        # Formato BR com pontos de milhar: "1.234,56" → "1234.56"
        if re.match(r"^\d{1,3}(\.\d{3})+(,\d+)?$", s):
            s = s.replace(".", "").replace(",", ".")
        else:
            # Remove separador de milhar (vírgula no formato US: "1,234.56")
            s = s.replace(",", "")
        try:
            return Decimal(s)
        except InvalidOperation:
            return None
