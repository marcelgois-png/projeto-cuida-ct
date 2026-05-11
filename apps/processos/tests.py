from __future__ import annotations

import io
from datetime import date

from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import Workbook, load_workbook

from .forms import InteressadoProcessoFormSet, ProcessoForm
from .importers import gerar_modelo_xlsx, ProcessoImporter
from .models import InteressadoProcesso, Processo
from apps.tracker.models import AcompanhamentoRequisicao, EncaminhamentoDiretor, Requisicao, StatusRequisicao


User = get_user_model()


class ProcessoImporterTests(TestCase):
    def _xlsx_upload(self, wb: Workbook) -> SimpleUploadedFile:
        buf = io.BytesIO()
        wb.save(buf)
        return SimpleUploadedFile(
            "processos.xlsx",
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_modelo_inclui_unidade_origem_na_coluna_r_e_aba_interessados(self):
        wb = load_workbook(io.BytesIO(gerar_modelo_xlsx()), read_only=True, data_only=True)

        self.assertEqual(wb["PROCESSOS"]["R1"].value, "unidade_origem")
        self.assertIn("INTERESSADOS_PROCESSO", wb.sheetnames)
        self.assertEqual(
            [wb["INTERESSADOS_PROCESSO"].cell(1, col).value for col in range(1, 5)],
            ["numero_processo", "tipo", "identificador", "nome"],
        )

    def test_importa_unidade_origem_e_interessados(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "PROCESSOS"
        ws.append([
            "numero_processo",
            "assunto",
            "data_abertura",
            "data_os",
            "data_conclusao",
            "data_arquivamento",
            "status",
            "situacao_sipac",
            "gerencia",
            "servico",
            "predio",
            "tipo_ambiente",
            "empresa",
            "classificacao_az",
            "link_sipac",
            "observacao",
            "acompanhamento_ct",
            "unidade_origem",
        ])
        ws.append([
            "23074.001234/2024-01",
            "Reforma do telhado",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "B",
            "https://sipac.ufpb.br/public/jsp/processos/processo_detalhado.jsf?id=1",
            "",
            "",
            "CT - DIREÇÃO DE CENTRO (11.01.17.01)",
        ])

        wi = wb.create_sheet("INTERESSADOS_PROCESSO")
        wi.append(["numero_processo", "tipo", "identificador", "nome"])
        interessados = [
            ("Unidade", "110055", "CENTRO DE TECNOLOGIA (CT)"),
            ("Servidor", "333695", "EUGENIO CORTE REAL COUTINHO"),
            ("Servidor", "1630306", "MARCEL DE GOIS PINTO"),
            ("Servidor", "1529493", "MARCO ANTONIO FARIAS COUTINHO"),
            ("Servidor", "1474918", "NIVALDO VIEIRA DO NASCIMENTO JUNIOR"),
        ]
        for interessado in interessados:
            wi.append(["23074.001234/2024-01", *interessado])

        resultado = ProcessoImporter().import_file(self._xlsx_upload(wb))

        processo = Processo.objects.get(numero_processo="23074.001234/2024-01")
        self.assertEqual(resultado["criados"], 1)
        self.assertEqual(resultado["interessados_importados"], 5)
        self.assertEqual(processo.unidade_origem, "CT - DIREÇÃO DE CENTRO (11.01.17.01)")
        self.assertEqual(
            list(
                InteressadoProcesso.objects
                .filter(processo=processo)
                .order_by("tipo", "identificador")
                .values_list("tipo", "identificador", "nome")
            ),
            [
                ("Servidor", "1474918", "NIVALDO VIEIRA DO NASCIMENTO JUNIOR"),
                ("Servidor", "1529493", "MARCO ANTONIO FARIAS COUTINHO"),
                ("Servidor", "1630306", "MARCEL DE GOIS PINTO"),
                ("Servidor", "333695", "EUGENIO CORTE REAL COUTINHO"),
                ("Unidade", "110055", "CENTRO DE TECNOLOGIA (CT)"),
            ],
        )

    def test_formulario_salva_multiplos_interessados_estruturados(self):
        form = ProcessoForm(data={
            "numero_processo": "23074.009999/2024-01",
            "assunto": "Teste de cadastro",
            "classificacao_az": "A",
            "unidade_origem": "CT - DIREÇÃO DE CENTRO (11.01.17.01)",
        })
        prefix = InteressadoProcessoFormSet().prefix
        formset = InteressadoProcessoFormSet(data={
            f"{prefix}-TOTAL_FORMS": "2",
            f"{prefix}-INITIAL_FORMS": "0",
            f"{prefix}-MIN_NUM_FORMS": "0",
            f"{prefix}-MAX_NUM_FORMS": "1000",
            f"{prefix}-0-tipo": "Servidor",
            f"{prefix}-0-identificador": "333695",
            f"{prefix}-0-nome": "EUGENIO CORTE REAL COUTINHO",
            f"{prefix}-1-tipo": "Servidor",
            f"{prefix}-1-identificador": "1630306",
            f"{prefix}-1-nome": "MARCEL DE GOIS PINTO",
        })

        self.assertTrue(form.is_valid(), form.errors)
        self.assertTrue(formset.is_valid(), formset.errors)
        processo = form.save()
        formset.instance = processo
        formset.save()

        self.assertEqual(processo.unidade_origem, "CT - DIREÇÃO DE CENTRO (11.01.17.01)")
        self.assertEqual(
            list(processo.interessados.order_by("identificador").values_list("tipo", "identificador", "nome")),
            [
                ("Servidor", "1630306", "MARCEL DE GOIS PINTO"),
                ("Servidor", "333695", "EUGENIO CORTE REAL COUTINHO"),
            ],
        )


class ProcessoCadastroFromEncaminhamentoTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="operador", password="segredo", role="operator")
        self.client.login(username="operador", password="segredo")
        status, _ = StatusRequisicao.objects.get_or_create(
            codigo="02 ENVIADA",
            defaults={"nome": "Enviada", "mapeamento_situacao": "ATIVA", "ordem": 2},
        )
        self.requisicao = Requisicao.objects.create(
            codigo="3198/2026",
            numero=3198,
            ano=2026,
            assunto="Desinstalacao de aparelho de ar-condicionado",
            data_cadastro=date(2026, 4, 8),
            status_sipac=status,
            local_servico="Bloco CTKLM",
            nome_requisitante_snapshot="LISZANDRA FERNANDA ARAUJO CAMPOS",
        )
        self.encaminhamento = EncaminhamentoDiretor.objects.create(
            tipo=EncaminhamentoDiretor.Tipo.ABRIR_PROCESSO,
            orientacoes="Abrir processo administrativo.",
            diretor=self.user,
        )
        self.encaminhamento.requisicoes.add(self.requisicao)

    def test_cadastro_vincula_processo_ao_encaminhamento_e_a_requisicao(self):
        prefix = InteressadoProcessoFormSet().prefix
        response = self.client.post(
            f"/processos/cadastro/novo/?encaminhamento={self.encaminhamento.pk}",
            {
                "encaminhamento_diretor": str(self.encaminhamento.pk),
                "numero_processo": "23074.003198/2026-01",
                "assunto": "Processo aberto a partir de requisicao",
                "classificacao_az": "A",
                f"{prefix}-TOTAL_FORMS": "0",
                f"{prefix}-INITIAL_FORMS": "0",
                f"{prefix}-MIN_NUM_FORMS": "0",
                f"{prefix}-MAX_NUM_FORMS": "1000",
            },
        )

        self.assertEqual(response.status_code, 302)

        processo = Processo.objects.get(numero_processo="23074.003198/2026-01")
        self.assertEqual(processo.encaminhamento_diretor, self.encaminhamento)
        self.assertEqual(list(processo.requisicoes.all()), [self.requisicao])
        self.assertEqual(list(self.requisicao.processos.all()), [processo])
        self.requisicao.refresh_from_db()
        self.assertEqual(self.requisicao.status_sipac.codigo, "Requisição Convertida em Processo")
        self.assertEqual(self.requisicao.situacao_requisicao, "Inativa")
        self.assertTrue(
            AcompanhamentoRequisicao.objects.filter(
                requisicao=self.requisicao,
                atualizacao_situacao=(
                    "Requisição convertida em processo devido à complexidade dos serviços. "
                    "Para acompanhar a continuidade do atendimento, consulte o processo "
                    "23074.003198/2026-01."
                ),
            ).exists()
        )
