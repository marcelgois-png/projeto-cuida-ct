from __future__ import annotations

import csv
import io
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .domain import (
    clean_display_text,
    coerce_brazilian_phone,
    coerce_coordinate,
    coerce_date,
    coerce_int,
    derive_situation,
    extract_request_parts,
    normalize_key,
    normalize_text,
    resolve_status_sipac_metadata,
)
from .models import (
    DivisaoSINFRA,
    HistoricoStatus,
    ImportacaoArquivo,
    Predio,
    RegraPrioridade,
    Requisicao,
    Servico,
    Solicitante,
    StatusRequisicao,
    TaxonomiaServico,
    TipoServico,
)
from .services import register_status_history, resolve_priority_label


class ImportErrorPlanilha(Exception):
    pass


# Modelo de cadastro em lote: chave normalizada -> rótulo exibido na planilha.
# Apenas as colunas obrigatórias são exigidas; as opcionais podem ficar em branco.
CADASTRO_LOTE_COLUNAS_OBRIGATORIAS: dict[str, str] = {
    "no requisicao": "Nº Requisição",
    "assunto": "Assunto",
    "data de cadastro": "Data de Cadastro",
    "divisao": "Divisão",
    "tipo de servico": "Tipo de Serviço",
    "servico": "Serviço",
    "status sipac": "Status SIPAC",
    "predio envolvido": "Prédio Envolvido",
    "requisitante": "Requisitante",
    "unidade setor": "Unidade/Setor",
}

CADASTRO_LOTE_COLUNAS_OPCIONAIS: dict[str, str] = {
    "orcamento": "Orçamento",
    "data execucao": "Data de Execução",
    "local do servico": "Local do Serviço",
    "contato": "Contato",
    "link do atendimento": "Link do Atendimento",
    "link sipac": "Link SIPAC",
}

# Ordem das colunas geradas no modelo XLSX para download (obrigatórias + opcionais).
CADASTRO_LOTE_COLUNAS: dict[str, str] = {
    **CADASTRO_LOTE_COLUNAS_OBRIGATORIAS,
    **CADASTRO_LOTE_COLUNAS_OPCIONAIS,
}


class WorkbookImporter:
    def __init__(self, user=None):
        self.user = user
        # Caches por execução de importação: evitam refazer o mesmo
        # get_or_create/lookup para valores repetidos linha após linha.
        self._cache_predio: dict[str, Predio | None] = {}
        self._cache_solicitante: dict[str, Solicitante | None] = {}
        self._cache_servico_fks: dict[
            tuple[str, str, str],
            tuple[DivisaoSINFRA | None, TipoServico | None, Servico | None],
        ] = {}
        self._cache_status: dict[str, StatusRequisicao | None] = {}
        self._cache_priority_rule: dict[str, str] = {}

    def import_file(self, uploaded_file) -> ImportacaoArquivo:
        suffix = Path(uploaded_file.name).suffix.lower()
        if suffix not in {".xlsm", ".xlsx", ".csv"}:
            raise ImportErrorPlanilha("Formato inválido. Envie um arquivo XLSM, XLSX ou CSV.")

        importacao = ImportacaoArquivo.objects.create(
            nome_arquivo=uploaded_file.name,
            tipo_arquivo=suffix.lstrip("."),
            status=ImportacaoArquivo.Status.PROCESSANDO,
            iniciado_por=self.user,
        )

        try:
            with transaction.atomic():
                if suffix == ".csv":
                    summary = self._import_csv(uploaded_file, importacao)
                else:
                    summary = self._import_workbook(uploaded_file, importacao, suffix)
            importacao.status = ImportacaoArquivo.Status.CONCLUIDA
            importacao.resumo_json = summary
            importacao.processado_em = timezone.now()
            importacao.save(update_fields=["status", "resumo_json", "processado_em", "atualizado_em"])
        except ImportErrorPlanilha as exc:
            self._registrar_falha_importacao(importacao, str(exc))
            raise
        except Exception as exc:  # noqa: BLE001 - registra a causa antes de propagar
            detalhe = f"{type(exc).__name__}: {exc}".strip()
            self._registrar_falha_importacao(
                importacao, f"Erro inesperado na importação. {detalhe}"
            )
            raise ImportErrorPlanilha(
                "Não foi possível processar o arquivo. "
                f"Verifique se ele segue o formato esperado. Detalhe técnico: {detalhe}"
            ) from exc

        return importacao

    def _import_workbook(self, uploaded_file, importacao: ImportacaoArquivo, suffix: str) -> dict[str, Any]:
        uploaded_file.seek(0)
        workbook = load_workbook(
            uploaded_file,
            read_only=True,
            data_only=True,
            keep_vba=suffix == ".xlsm",
        )
        requisitantes_lookup = self._load_requisitantes(workbook)
        self._load_predios(workbook)
        self._load_status_sipac_options(workbook)
        self._load_priority_rules(workbook)
        self._load_taxonomy_candidates(workbook)
        references = self._load_reference_priority_maps(workbook)
        base_sheet = self._get_sheet(workbook, "requisicoes - ct")
        rows = self._sheet_to_dicts(base_sheet)
        summary = self._upsert_requests(rows, importacao, requisitantes_lookup, references)
        summary["referencias"] = {
            "prioridades_gme": len(references["gme"]),
            "prioridades_ar": len(references["ar"]),
            "requisitantes": len(requisitantes_lookup),
        }
        return summary

    def _import_csv(self, uploaded_file, importacao: ImportacaoArquivo) -> dict[str, Any]:
        uploaded_file.seek(0)
        text = self._decode_csv_bytes(uploaded_file.read())
        reader = csv.DictReader(io.StringIO(text), delimiter=self._sniff_delimiter(text))
        rows = []
        for row in reader:
            rows.append({self._normalized_header(key): value for key, value in row.items()})
        return self._upsert_requests(rows, importacao, {}, {"gme": {}, "ar": {}})

    def _decode_csv_bytes(self, raw: bytes) -> str:
        # CSVs do Excel/SIPAC em pt-BR costumam vir em Windows-1252, não UTF-8.
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                return raw.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise ImportErrorPlanilha(
            "Não foi possível ler o arquivo CSV: codificação de caracteres não reconhecida. "
            "Salve o arquivo como UTF-8 ou Windows-1252 e tente novamente."
        )

    def _sniff_delimiter(self, text: str) -> str:
        sample = text[:4096]
        first_line = sample.splitlines()[0] if sample else ""
        if first_line.count(";") > first_line.count(","):
            return ";"
        return ","

    def _upsert_requests(
        self,
        rows: list[dict[str, Any]],
        importacao: ImportacaoArquivo,
        requisitantes_lookup: dict[str, dict[str, str]],
        references: dict[str, dict[str, dict[str, str]]],
    ) -> dict[str, Any]:
        created = 0
        updated = 0
        counters = {
            "divisao": Counter(),
            "status_sipac": Counter(),
            "situacao_requisicao": Counter(),
            "sinfra_responsavel": Counter(),
            "prioridade_final": Counter(),
        }

        history_to_create: list[HistoricoStatus] = []
        for row in rows:
            codigo = self._get_value(row, "n requisicao", "no requisicao", "codigo")
            if not codigo:
                continue
            numero, ano = extract_request_parts(str(codigo))
            requisitante_nome = self._get_value(row, "requisitante")
            requisitante_data = requisitantes_lookup.get(normalize_key(requisitante_nome), {})
            unidade_setor = self._get_value(row, "unidade/setor", "unidade setor") or requisitante_data.get("setor", "")
            contato = coerce_brazilian_phone(self._get_value(row, "contato") or requisitante_data.get("contato", ""))

            predio = self._ensure_predio(self._get_value(row, "predio envolvido"))
            solicitante = self._ensure_solicitante(requisitante_nome, contato)
            reference = references["gme"].get(str(codigo)) or references["ar"].get(str(codigo)) or {}

            divisao_str = clean_display_text(self._get_value(row, "divisao") or "")
            tipo_str = clean_display_text(self._get_value(row, "tipo de servico") or "")
            servico_str = clean_display_text(self._get_value(row, "servico") or "")
            status_str = clean_display_text(self._get_value(row, "status sipac") or "")
            divisao_fk, tipo_fk, servico_fk = self._resolve_service_fks(divisao_str, tipo_str, servico_str)
            status_fk = self._resolve_status_fk(status_str)

            defaults = {
                "numero": numero,
                "ano": ano,
                "assunto": clean_display_text(self._get_value(row, "assunto") or ""),
                "orcamento": self._get_value(row, "orcamento", "orçamento") or "",
                "data_cadastro": coerce_date(self._get_value(row, "data de cadastro")),
                "tipo_requisicao": clean_display_text(self._get_value(row, "tipo de requisicao") or ""),
                "divisao": divisao_fk,
                "unidade_origem": clean_display_text(self._get_value(row, "unidade de origem") or ""),
                "status_sipac": status_fk,
                "tipo_servico": tipo_fk,
                "servico": servico_fk,
                "predio": predio,
                "local_servico": clean_display_text(self._get_value(row, "local do servico") or ""),
                "latitude": coerce_coordinate(self._get_value(row, "latitude"), kind="latitude"),
                "longitude": coerce_coordinate(self._get_value(row, "longitude"), kind="longitude"),
                "solicitante": solicitante,
                "nome_requisitante_snapshot": clean_display_text(requisitante_nome or ""),
                "unidade_setor_snapshot": clean_display_text(unidade_setor),
                "contato_direto_url": contato or "",
                "situacao_texto": clean_display_text(self._get_value(row, "situacao") or reference.get("situacao", "") or ""),
                "status_fluxo": clean_display_text(self._get_value(row, "status") or ""),
                "gravidade": clean_display_text(self._get_value(row, "gravidade") or ""),
                "urgencia": clean_display_text(self._get_value(row, "urgencia") or ""),
                "tendencia": clean_display_text(self._get_value(row, "tendencia") or ""),
                "sinfra_responsavel": clean_display_text(self._get_value(row, "sinfra") or reference.get("sinfra", "") or ""),
                "link_atendimento": clean_display_text(self._get_value(row, "link do atendimento") or reference.get("link_atendimento", "") or ""),
                "link_sipac": clean_display_text(self._get_value(row, "link sipac") or reference.get("link_sipac", "") or ""),
                "data_execucao": coerce_date(self._get_value(row, "data execucao", "data de execucao")),
                "dias_para_execucao": coerce_int(self._get_value(row, "dias para execucao")),
                "visivel_publicamente": True,
                "importacao_origem": importacao,
            }

            priority_key = normalize_key(
                divisao_fk.nome if divisao_fk else "",
                tipo_fk.nome if tipo_fk else "",
                servico_fk.nome if servico_fk else "",
            )
            priority_from_rule = self._cached_priority_from_rule(priority_key)
            defaults["prioridade_final"] = (
                reference.get("prioridade_final") or priority_from_rule or ""
            )

            requisicao, was_created = Requisicao.objects.get_or_create(
                codigo=str(codigo), defaults=defaults
            )
            previous_status = (
                requisicao.status_sipac.codigo if (not was_created and requisicao.status_sipac) else ""
            )
            previous_note = requisicao.situacao_texto if not was_created else ""
            if was_created:
                created += 1
            else:
                updated += 1
                prioridade_anterior = requisicao.prioridade_final
                for field, value in defaults.items():
                    setattr(requisicao, field, value)
                requisicao.prioridade_final = (
                    reference.get("prioridade_final")
                    or priority_from_rule
                    or prioridade_anterior
                )
                requisicao.save()
            if was_created:
                history_to_create.append(
                    HistoricoStatus(
                        requisicao=requisicao,
                        status_sipac=requisicao.status_sipac.codigo if requisicao.status_sipac else "",
                        situacao_requisicao=requisicao.situacao_requisicao,
                        observacao=requisicao.situacao_texto,
                        origem=HistoricoStatus.Origem.IMPORTACAO,
                        usuario=self.user,
                    )
                )
            else:
                register_status_history(
                    requisicao,
                    previous_status=previous_status,
                    previous_note=previous_note,
                    note=requisicao.situacao_texto,
                    user=self.user,
                    origin=HistoricoStatus.Origem.IMPORTACAO,
                )

            counters["divisao"][requisicao.divisao.nome if requisicao.divisao else ""] += 1
            counters["status_sipac"][requisicao.status_sipac.codigo if requisicao.status_sipac else ""] += 1
            counters["situacao_requisicao"][requisicao.situacao_requisicao] += 1
            counters["sinfra_responsavel"][requisicao.sinfra_responsavel] += 1
            counters["prioridade_final"][requisicao.prioridade_final] += 1

        if history_to_create:
            HistoricoStatus.objects.bulk_create(history_to_create, batch_size=500)

        return {
            "criados": created,
            "atualizados": updated,
            "total_processado": created + updated,
            "contagens": {key: dict(value) for key, value in counters.items()},
        }

    def _load_requisitantes(self, workbook) -> dict[str, dict[str, str]]:
        sheet = self._get_sheet(workbook, "requisitantes do ct")
        lookup: dict[str, dict[str, str]] = {}
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if index == 1:
                continue
            nome = normalize_text(row[0] if len(row) > 0 else "")
            if not nome:
                continue
            setor = normalize_text(row[1] if len(row) > 1 else "")
            contato = coerce_brazilian_phone(row[2] if len(row) > 2 else "")
            solicitante, _ = Solicitante.objects.get_or_create(
                chave_normalizada=normalize_key(nome),
                defaults={"nome": nome, "contato_url": contato},
            )
            changed = False
            if not solicitante.nome:
                solicitante.nome = nome
                changed = True
            if contato and solicitante.contato_url != contato:
                solicitante.contato_url = contato
                changed = True
            if changed:
                solicitante.save()
            lookup[normalize_key(nome)] = {"setor": setor, "contato": contato}
        return lookup

    def _load_predios(self, workbook) -> None:
        sheet = self._get_sheet(workbook, "legendas (1)")
        for row in sheet.iter_rows(min_row=2, max_col=16, values_only=True):
            nome = normalize_text(row[13] if len(row) > 13 else "")
            if not nome:
                continue
            latitude = coerce_coordinate(row[14] if len(row) > 14 else None, kind="latitude")
            longitude = coerce_coordinate(row[15] if len(row) > 15 else None, kind="longitude")
            predio, created = Predio.objects.get_or_create(
                chave_normalizada=normalize_key(nome),
                defaults={"nome": nome, "latitude": latitude, "longitude": longitude},
            )
            if not created:
                changed = False
                if not predio.nome:
                    predio.nome = nome
                    changed = True
                if latitude is not None and predio.latitude != latitude:
                    predio.latitude = latitude
                    changed = True
                if longitude is not None and predio.longitude != longitude:
                    predio.longitude = longitude
                    changed = True
                if changed:
                    predio.save()

    def _load_status_sipac_options(self, workbook) -> None:
        sheet = self._get_sheet(workbook, "legendas (1)")
        seen: dict[str, int] = {}
        for row in sheet.iter_rows(min_row=2, min_col=23, max_col=23, values_only=True):
            descricao = clean_display_text(row[0] if row else "")
            if not descricao:
                continue
            metadata = resolve_status_sipac_metadata(descricao)
            codigo = metadata["descricao"] or descricao
            ordem = metadata["ordem"] or seen.setdefault(descricao, len(seen) + 1)
            StatusRequisicao.objects.update_or_create(
                codigo=codigo,
                defaults={
                    "numero": metadata["numero"],
                    "nome": metadata["rotulo"] or descricao,
                    "chave_normalizada": normalize_key(codigo),
                    "ordem": ordem,
                    "ativa": True,
                    "mapeamento_situacao": "ATIVA" if derive_situation(codigo) == "Ativa" else "INATIVA",
                },
            )

    def _load_priority_rules(self, workbook) -> None:
        sheet = self._get_sheet(workbook, "legendas (1)")
        for row in sheet.iter_rows(min_row=2, min_col=44, max_col=45, values_only=True):
            key = normalize_key(row[0] if len(row) > 0 else "")
            priority = normalize_text(row[1] if len(row) > 1 else "")
            if not key or not priority:
                continue
            RegraPrioridade.objects.update_or_create(
                chave_normalizada=key,
                defaults={"prioridade": priority, "descricao": "Importada da planilha"},
            )

    def _load_taxonomy_candidates(self, workbook) -> None:
        if self._load_taxonomy_candidates_from_legendas_1(workbook):
            return
        self._load_taxonomy_candidates_from_legendas_2(workbook)

    def _load_taxonomy_candidates_from_legendas_1(self, workbook) -> bool:
        sheet = self._get_sheet(workbook, "legendas (1)")
        divisao_orders: dict[str, int] = {}
        tipo_orders: dict[tuple[str, str], int] = {}
        servico_orders: dict[tuple[str, str, str], int] = {}
        found = False

        for row in sheet.iter_rows(min_row=2, min_col=27, max_col=29, values_only=True):
            divisao = self._clean_display_text(row[0] if len(row) > 0 else "")
            tipo = self._clean_display_text(row[1] if len(row) > 1 else "")
            servico = self._clean_display_text(row[2] if len(row) > 2 else "")
            if not divisao and not tipo and not servico:
                continue
            found = True
            ordem_divisao = divisao_orders.setdefault(divisao, len(divisao_orders) + 1) if divisao else None
            ordem_tipo = None
            ordem_servico = None
            if divisao and tipo:
                ordem_tipo = tipo_orders.setdefault((divisao, tipo), len([key for key in tipo_orders if key[0] == divisao]) + 1)
            if divisao and tipo and servico:
                ordem_servico = servico_orders.setdefault(
                    (divisao, tipo, servico),
                    len([key for key in servico_orders if key[0] == divisao and key[1] == tipo]) + 1,
                )
            self._upsert_taxonomia_hierarchy(divisao, tipo, servico, ordem_divisao, ordem_tipo, ordem_servico)

        return found

    def _load_taxonomy_candidates_from_legendas_2(self, workbook) -> None:
        sheet = self._get_sheet(workbook, "legendas (2)")
        division_types: dict[str, list[str]] = defaultdict(list)
        divisao_orders: dict[str, int] = {}
        tipo_orders: dict[tuple[str, str], int] = {}
        servico_orders: dict[tuple[str, str, str], int] = {}

        for row in sheet.iter_rows(min_row=2, min_col=18, max_col=25, values_only=True):
            divisao = self._clean_display_text(row[0] if len(row) > 0 else "")
            if not divisao:
                continue
            ordem_divisao = divisao_orders.setdefault(divisao, len(divisao_orders) + 1)
            self._upsert_taxonomia_hierarchy(divisao, "", "", ordem_divisao, None, None)
            for index, tipo in enumerate(row[1:], start=1):
                tipo_text = self._clean_display_text(tipo)
                if not tipo_text:
                    continue
                division_types[divisao].append(tipo_text)
                tipo_orders.setdefault((divisao, tipo_text), index)
                self._upsert_taxonomia_hierarchy(divisao, tipo_text, "", ordem_divisao, index, None)

        for row in sheet.iter_rows(min_row=2, min_col=27, max_col=37, values_only=True):
            tipo = self._clean_display_text(row[0] if len(row) > 0 else "")
            if not tipo:
                continue
            divisao = next((key for key, values in division_types.items() if tipo in values), "")
            ordem_divisao = divisao_orders.get(divisao)
            ordem_tipo = tipo_orders.get((divisao, tipo))
            for index, servico in enumerate(row[1:], start=1):
                servico_text = self._clean_display_text(servico)
                if not servico_text or not divisao:
                    continue
                servico_orders.setdefault((divisao, tipo, servico_text), index)
                self._upsert_taxonomia_hierarchy(divisao, tipo, servico_text, ordem_divisao, ordem_tipo, index)

    def _upsert_taxonomia_hierarchy(
        self,
        divisao: str,
        tipo_servico: str,
        servico: str,
        ordem_divisao: int | None,
        ordem_tipo: int | None,
        ordem_servico: int | None,
    ) -> None:
        if divisao:
            TaxonomiaServico.objects.update_or_create(
                chave_normalizada=normalize_key(divisao),
                defaults={
                    "divisao": divisao,
                    "tipo_servico": "",
                    "servico": "",
                    "ordem_divisao": ordem_divisao,
                    "ordem_tipo": None,
                    "ordem_servico": None,
                },
            )
        if divisao and tipo_servico:
            TaxonomiaServico.objects.update_or_create(
                chave_normalizada=normalize_key(divisao, tipo_servico),
                defaults={
                    "divisao": divisao,
                    "tipo_servico": tipo_servico,
                    "servico": "",
                    "ordem_divisao": ordem_divisao,
                    "ordem_tipo": ordem_tipo,
                    "ordem_servico": None,
                },
            )
        if divisao and tipo_servico and servico:
            TaxonomiaServico.objects.update_or_create(
                chave_normalizada=normalize_key(divisao, tipo_servico, servico),
                defaults={
                    "divisao": divisao,
                    "tipo_servico": tipo_servico,
                    "servico": servico,
                    "ordem_divisao": ordem_divisao,
                    "ordem_tipo": ordem_tipo,
                    "ordem_servico": ordem_servico,
                },
            )

    def _load_reference_priority_maps(self, workbook) -> dict[str, dict[str, dict[str, str]]]:
        return {
            "gme": self._load_reference_sheet(workbook, "prioridades sinfra - gme"),
            "ar": self._load_reference_sheet(workbook, "prioridades ar-condicionado"),
        }

    def _load_reference_sheet(self, workbook, sheet_name: str) -> dict[str, dict[str, str]]:
        sheet = self._get_sheet(workbook, sheet_name)
        references = {}
        for row in self._sheet_to_dicts(sheet):
            codigo = self._get_value(row, "n requisicao", "no requisicao")
            if not codigo:
                continue
            references[str(codigo)] = {
                "sinfra": clean_display_text(self._get_value(row, "sinfra")),
                "prioridade_final": clean_display_text(self._get_value(row, "prioridade")),
                "link_atendimento": clean_display_text(self._get_value(row, "link do atendimento")),
                "link_sipac": clean_display_text(self._get_value(row, "link sipac")),
                "situacao": clean_display_text(self._get_value(row, "situacao")),
            }
        return references

    def _ensure_predio(self, nome: str | None) -> Predio | None:
        if not nome:
            return None
        normalized = normalize_key(nome)
        if normalized in self._cache_predio:
            return self._cache_predio[normalized]
        predio = Predio.objects.filter(chave_normalizada=normalized).first()
        if not predio:
            predio = Predio.objects.create(nome=normalize_text(nome))
        self._cache_predio[normalized] = predio
        return predio

    def _ensure_solicitante(self, nome: str | None, contato: str) -> Solicitante | None:
        if not nome:
            return None
        normalized = normalize_key(nome)
        if normalized in self._cache_solicitante:
            return self._cache_solicitante[normalized]
        solicitante = Solicitante.objects.filter(chave_normalizada=normalized).first()
        if solicitante:
            if contato and solicitante.contato_url != contato:
                solicitante.contato_url = contato
                solicitante.save()
        else:
            solicitante = Solicitante.objects.create(
                nome=normalize_text(nome),
                contato_url=coerce_brazilian_phone(contato),
            )
        self._cache_solicitante[normalized] = solicitante
        return solicitante

    def _cached_priority_from_rule(self, priority_key: str) -> str:
        if not priority_key:
            return ""
        if priority_key in self._cache_priority_rule:
            return self._cache_priority_rule[priority_key]
        rule = RegraPrioridade.objects.filter(
            chave_normalizada=priority_key, ativa=True
        ).first()
        priority = rule.prioridade if rule else ""
        self._cache_priority_rule[priority_key] = priority
        return priority

    def _resolve_status_fk(self, status_str: str) -> StatusRequisicao | None:
        if not status_str:
            return None
        if status_str in self._cache_status:
            return self._cache_status[status_str]
        status_fk = StatusRequisicao.objects.filter(codigo=status_str).first()
        self._cache_status[status_str] = status_fk
        return status_fk

    def _resolve_service_fks(
        self, divisao: str, tipo_servico: str, servico: str
    ) -> tuple[DivisaoSINFRA | None, TipoServico | None, Servico | None]:
        if not divisao:
            return None, None, None
        cache_key = (divisao, tipo_servico, servico)
        if cache_key in self._cache_servico_fks:
            return self._cache_servico_fks[cache_key]
        divisao_fk, _ = DivisaoSINFRA.objects.get_or_create(nome=divisao)
        tipo_fk = None
        servico_fk = None
        if tipo_servico:
            tipo_fk, _ = TipoServico.objects.get_or_create(nome=tipo_servico, divisao=divisao_fk)
        if servico and tipo_fk:
            servico_fk, _ = Servico.objects.get_or_create(nome=servico, tipo_servico=tipo_fk)
        result = (divisao_fk, tipo_fk, servico_fk)
        self._cache_servico_fks[cache_key] = result
        return result

    def _ensure_taxonomia(self, divisao: str | None, tipo_servico: str | None, servico: str | None) -> TaxonomiaServico | None:
        if not divisao and not tipo_servico and not servico:
            return None
        divisao = self._clean_display_text(divisao)
        tipo_servico = self._clean_display_text(tipo_servico)
        servico = self._clean_display_text(servico)
        key = normalize_key(divisao, tipo_servico, servico)
        taxonomia = TaxonomiaServico.objects.filter(chave_normalizada=key).first()
        if taxonomia:
            return taxonomia
        return TaxonomiaServico.objects.create(
            divisao=divisao,
            tipo_servico=tipo_servico,
            servico=servico,
        )

    def _get_sheet(self, workbook, normalized_title: str):
        for sheet in workbook.worksheets:
            if self._normalized_header(sheet.title) == self._normalized_header(normalized_title):
                return sheet
        raise ImportErrorPlanilha(f"Aba obrigatória não encontrada: {normalized_title}")

    def _sheet_to_dicts(self, sheet) -> list[dict[str, Any]]:
        iterator = sheet.iter_rows(values_only=True)
        try:
            header = next(iterator)
        except StopIteration as exc:
            raise ImportErrorPlanilha(f"Aba vazia: {sheet.title}") from exc
        header_map = [self._normalized_header(column) for column in header]
        rows = []
        for raw_row in iterator:
            row = {}
            for index, column in enumerate(header_map):
                if not column:
                    continue
                row[column] = raw_row[index] if index < len(raw_row) else None
            rows.append(row)
        return rows

    def _get_value(self, row: dict[str, Any], *aliases: str) -> Any:
        for alias in aliases:
            key = self._normalized_header(alias)
            if key in row:
                return row[key]
        return None

    def _normalized_header(self, value: Any) -> str:
        normalized = normalize_text(value).lower()
        normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
        return normalized.strip()

    def _clean_display_text(self, value: Any) -> str:
        return clean_display_text(value)

    # ------------------------------------------------------------------
    # Cadastro em lote (modelo XLSX enxuto)
    # ------------------------------------------------------------------
    def build_modelo_cadastro_lote(self) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Requisições"
        sheet.append(list(CADASTRO_LOTE_COLUNAS.values()))
        sheet.append(
            [
                "123/2026",
                "Troca de lâmpada queimada na sala 10",
                "19/05/2026",
                "Elétrica",
                "Iluminação",
                "Troca de lâmpada",
                "01 ABERTA",
                "Bloco CT",
                "João Silva",
                "Coordenação de Exemplo",
                "1.500,00",
                "26/05/2026",
                "Sala 10, 1º andar",
                "(83) 99999-9999",
                "https://exemplo.org/atendimento/123",
                "https://sipac.ufpb.br/requisicao/123",
            ]
        )
        for index in range(1, len(CADASTRO_LOTE_COLUNAS) + 1):
            sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = 26

        instrucoes = workbook.create_sheet("Instruções")
        for linha in (
            ["Como usar este modelo de cadastro em lote"],
            [""],
            ["1. Preencha uma requisição por linha, a partir da linha 2 da aba 'Requisições'."],
            ["2. A linha 2 traz um exemplo: apague-a antes de enviar (ou substitua pelos seus dados)."],
            ["3. Não altere, renomeie nem remova as colunas do cabeçalho."],
            [
                "4. Campos obrigatórios: "
                + ", ".join(CADASTRO_LOTE_COLUNAS_OBRIGATORIAS.values())
                + "."
            ],
            [
                "5. Campos opcionais (podem ficar em branco): "
                + ", ".join(CADASTRO_LOTE_COLUNAS_OPCIONAIS.values())
                + "."
            ],
            ["6. Data de Cadastro e Data de Execução no formato dd/mm/aaaa (ex.: 19/05/2026)."],
            ["7. Nº Requisição no formato número/ano (ex.: 123/2026)."],
            ["8. Salve no formato XLSX (.xlsx) antes de enviar."],
            [""],
            ["Linhas com erro são informadas após o envio e podem ser corrigidas e reenviadas."],
        ):
            instrucoes.append(linha)
        instrucoes.column_dimensions["A"].width = 90
        return workbook

    def _registrar_falha_importacao(
        self, importacao: ImportacaoArquivo, mensagem: str
    ) -> None:
        importacao.status = ImportacaoArquivo.Status.FALHA
        importacao.mensagem_erro = mensagem[:2000]
        importacao.processado_em = timezone.now()
        importacao.save(
            update_fields=["status", "mensagem_erro", "processado_em", "atualizado_em"]
        )

    def import_cadastro_lote(self, uploaded_file) -> ImportacaoArquivo:
        suffix = Path(uploaded_file.name).suffix.lower()
        if suffix != ".xlsx":
            raise ImportErrorPlanilha(
                "Formato inválido para cadastro em lote. Baixe o modelo e envie o arquivo no formato XLSX (.xlsx)."
            )

        importacao = ImportacaoArquivo.objects.create(
            nome_arquivo=uploaded_file.name,
            tipo_arquivo="xlsx",
            status=ImportacaoArquivo.Status.PROCESSANDO,
            iniciado_por=self.user,
        )

        try:
            valid_rows, row_errors = self._read_cadastro_lote_rows(uploaded_file)
            summary: dict[str, Any] = {
                "criados": 0,
                "atualizados": 0,
                "total_processado": 0,
                "contagens": {},
            }
            if valid_rows:
                with transaction.atomic():
                    summary = self._upsert_requests(
                        valid_rows, importacao, {}, {"gme": {}, "ar": {}}
                    )
            summary["linhas_com_erro"] = len(row_errors)
            summary["erros_linhas"] = row_errors
        except ImportErrorPlanilha as exc:
            self._registrar_falha_importacao(importacao, str(exc))
            raise
        except Exception as exc:  # noqa: BLE001 - registra a causa antes de propagar
            detalhe = f"{type(exc).__name__}: {exc}".strip()
            self._registrar_falha_importacao(
                importacao, f"Erro inesperado no cadastro em lote. {detalhe}"
            )
            raise ImportErrorPlanilha(
                "Não foi possível concluir o cadastro em lote. "
                f"Verifique o arquivo e tente novamente. Detalhe técnico: {detalhe}"
            ) from exc

        if not valid_rows and row_errors:
            importacao.status = ImportacaoArquivo.Status.FALHA
            importacao.mensagem_erro = (
                f"Nenhuma linha pôde ser importada: {len(row_errors)} linha(s) com erro."
            )
        else:
            importacao.status = ImportacaoArquivo.Status.CONCLUIDA
        importacao.resumo_json = summary
        importacao.processado_em = timezone.now()
        importacao.save(
            update_fields=[
                "status",
                "mensagem_erro",
                "resumo_json",
                "processado_em",
                "atualizado_em",
            ]
        )
        return importacao

    def _read_cadastro_lote_rows(
        self, uploaded_file
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        uploaded_file.seek(0)
        try:
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        except Exception as exc:  # openpyxl levanta tipos variados p/ arquivo inválido
            raise ImportErrorPlanilha(
                "Não foi possível abrir o arquivo. Verifique se ele é um XLSX válido, "
                "gerado a partir do modelo de cadastro em lote, e não está corrompido."
            ) from exc

        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        try:
            header = next(iterator)
        except StopIteration as exc:
            raise ImportErrorPlanilha(
                "A planilha está vazia. Baixe o modelo, preencha as requisições e tente novamente."
            ) from exc

        header_map = [self._normalized_header(column) for column in header]
        present = {column for column in header_map if column}
        missing = [
            label
            for key, label in CADASTRO_LOTE_COLUNAS_OBRIGATORIAS.items()
            if key not in present
        ]
        if missing:
            raise ImportErrorPlanilha(
                "O arquivo não está no modelo esperado. Estão faltando as colunas: "
                + ", ".join(missing)
                + ". Baixe o modelo de cadastro em lote e use exatamente esse cabeçalho, "
                "sem renomear nem remover colunas."
            )

        valid_rows: list[dict[str, Any]] = []
        row_errors: list[dict[str, Any]] = []
        for line_no, raw_row in enumerate(iterator, start=2):
            row: dict[str, Any] = {}
            for index, column in enumerate(header_map):
                if not column:
                    continue
                row[column] = raw_row[index] if index < len(raw_row) else None
            if all(value is None or str(value).strip() == "" for value in row.values()):
                continue
            problems = self._validate_cadastro_lote_row(row)
            if problems:
                row_errors.append({"linha": line_no, "problemas": problems})
            else:
                valid_rows.append(row)
        return valid_rows, row_errors

    def _validate_cadastro_lote_row(self, row: dict[str, Any]) -> list[str]:
        problems: list[str] = []
        obrigatorios = {
            key: label
            for key, label in CADASTRO_LOTE_COLUNAS_OBRIGATORIAS.items()
            if key != "data de cadastro"
        }
        for key, label in obrigatorios.items():
            value = row.get(key)
            if value is None or str(value).strip() == "":
                problems.append(f"{label}: campo obrigatório não preenchido")

        codigo = row.get("no requisicao")
        if codigo and str(codigo).strip():
            numero, ano = extract_request_parts(str(codigo).strip())
            if numero is None or ano is None:
                problems.append(
                    "Nº Requisição: formato inválido (use número/ano, ex.: 123/2026)"
                )

        data = row.get("data de cadastro")
        if data is None or str(data).strip() == "":
            problems.append("Data de Cadastro: campo obrigatório não preenchido")
        elif coerce_date(data) is None:
            problems.append(
                "Data de Cadastro: data inválida (use dd/mm/aaaa, ex.: 19/05/2026)"
            )
        return problems
