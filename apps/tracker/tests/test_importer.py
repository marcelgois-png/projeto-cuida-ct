from datetime import date
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook

from apps.tracker.importers import (
    CADASTRO_LOTE_COLUNAS,
    CADASTRO_LOTE_COLUNAS_OBRIGATORIAS,
    ImportErrorPlanilha,
    WorkbookImporter,
)
from apps.tracker.models import Predio, RegraPrioridade, Requisicao, Solicitante, StatusRequisicao, TaxonomiaServico


class WorkbookImporterTests(TestCase):
    def build_workbook_file(self):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Requisições - CT"
        ws.append(
            [
                "Nº Requisição",
                "Assunto",
                "ORÇAMENTO",
                "Data de Cadastro",
                "Tipo de Requisição",
                "Divisão",
                "Unidade de Origem",
                "Status SIPAC",
                "Tipo de Serviço",
                "Serviço",
                "Prédio Envolvido",
                "Local do Serviço",
                "Dias deste abertura",
                "Requisitante",
                "Unidade/Setor",
                "Contato",
                "Situação",
                "Status",
                "GRAVIDADE",
                "URGÊNCIA",
                "TENDÊNCIA",
                "Link do Atendimento",
                "Link SIPAC",
                "SINFRA",
                "Data Execução",
                "Dias para execução",
            ]
        )
        ws.append(
            [
                "100/2026",
                "Troca de porta",
                "R$ 1.500,00",
                "01/03/2026",
                "REQUISIÇÃO DE MANUTENÇÃO",
                "Construção Civil",
                "CT - DIREÇÃO DE CENTRO",
                "04 OS EMITIDA",
                "Manutenção de Esquadrias",
                "Porta de Madeira",
                "Bloco A",
                "Sala 101",
                20,
                "MARIA TESTE",
                "CT - DIREÇÃO DE CENTRO",
                "https://wa.me/5583999999999",
                "Equipe acionada",
                "Em Progresso",
                "Grave",
                "É urgente",
                "Piorar em curto prazo",
                "",
                "https://sipac.local/requisicao/100",
                "",
                "",
                "",
            ]
        )
        ws.append(
            [
                "200/2026",
                "Vazamento em banheiro",
                "",
                "05/03/2026",
                "REQUISIÇÃO DE MANUTENÇÃO",
                "Instalações Hidráulicas e Sanitárias",
                "CT - DIREÇÃO DE CENTRO",
                "06 FINALIZADA",
                "Sanitária",
                "Desobstrução",
                "Bloco A",
                "Banheiro térreo",
                15,
                "JOÃO TESTE",
                "CT - DIREÇÃO DE CENTRO",
                "",
                "Concluído",
                "Concluído",
                "Pouco grave",
                "Pouco urgente",
                "Piorar em longo prazo",
                "",
                "https://sipac.local/requisicao/200",
                "",
                "10/03/2026",
                5,
            ]
        )

        requisitantes = workbook.create_sheet("Requisitantes do CT")
        requisitantes.append(["Nome", "Setor", "Contato"])
        requisitantes.append(["MARIA TESTE", "CT - DIREÇÃO DE CENTRO", "https://wa.me/5583999999999"])
        requisitantes.append(["JOÃO TESTE", "CT - DIREÇÃO DE CENTRO", ""])

        legendas1 = workbook.create_sheet("Legendas (1)")
        legendas1.append([""])

        linha_hidraulica = [None] * 45
        linha_hidraulica[13] = "Bloco A"
        linha_hidraulica[14] = -7123456
        linha_hidraulica[15] = -34654321
        linha_hidraulica[22] = "06 FINALIZADA"
        linha_hidraulica[26] = "Instalações Hidráulicas e Sanitárias"
        linha_hidraulica[27] = "Sanitária"
        linha_hidraulica[28] = "Desobstrução"
        linha_hidraulica[43] = "InstalacoesHidraulicaseSanitariasSanitariaDesobstrucao"
        linha_hidraulica[44] = "2 - Alta"
        legendas1.append(linha_hidraulica)

        linha_construcao = [None] * 45
        linha_construcao[22] = "04 OS EMITIDA"
        linha_construcao[26] = "Construção Civil"
        linha_construcao[27] = "Manutenção de Esquadrias"
        linha_construcao[28] = "Porta de Madeira"
        linha_construcao[43] = "ConstrucaoCivilManutencaodeEsquadriasPortadeMadeira"
        linha_construcao[44] = "1 - Urgente"
        legendas1.append(linha_construcao)

        legendas2 = workbook.create_sheet("Legendas (2)")
        legendas2.append([None] * 37)
        legendas2.append([None] * 17 + ["Construção Civil", "Manutenção de Esquadrias"])
        legendas2.append([None] * 17 + ["Instalações Hidráulicas e Sanitárias", "Sanitária"])
        legendas2.append([None] * 26 + ["Manutenção de Esquadrias", "Porta de Madeira"])
        legendas2.append([None] * 26 + ["Sanitária", "Desobstrução"])

        prioridades_gme = workbook.create_sheet("Prioridades SINFRA GME")
        prioridades_gme.append(["Nº Requisição", "Situação", "Link do Atendimento", "Link SIPAC", "SINFRA", "PRIORIDADE"])
        prioridades_gme.append(["100/2026", "Equipe acionada", "https://atendimento.local/100", "https://sipac.local/requisicao/100", "RN", "1 - Urgente"])

        prioridades_ar = workbook.create_sheet("Prioridades Ar-Condicionado")
        prioridades_ar.append(["Nº Requisição", "Situação", "Link do Atendimento", "Link SIPAC", "SINFRA", "PRIORIDADE"])
        prioridades_ar.append(["300/2026", "Em fila", "https://atendimento.local/300", "https://sipac.local/requisicao/300", "ECOAR", "2 - Alta"])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            "importacao.xlsx",
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_import_workbook_creates_data_and_preserves_priority_references(self):
        uploaded = self.build_workbook_file()
        result = WorkbookImporter().import_file(uploaded)

        self.assertEqual(result.resumo_json["total_processado"], 2)
        self.assertEqual(Requisicao.objects.count(), 2)
        self.assertEqual(Predio.objects.count(), 1)
        self.assertEqual(Solicitante.objects.count(), 2)
        self.assertGreaterEqual(TaxonomiaServico.objects.count(), 2)
        self.assertGreaterEqual(RegraPrioridade.objects.count(), 2)
        self.assertGreaterEqual(StatusRequisicao.objects.count(), 2)

        first = Requisicao.objects.get(codigo="100/2026")
        self.assertEqual(first.situacao_requisicao, "Ativa")
        self.assertEqual(first.prioridade_final, "1 - Urgente")
        self.assertEqual(first.sinfra_responsavel, "RN")
        self.assertEqual(first.link_atendimento, "https://atendimento.local/100")
        self.assertEqual(first.orcamento, "R$ 1.500,00")
        self.assertEqual(first.predio.nome, "Bloco A")
        self.assertEqual(first.divisao.nome, "Construção Civil")
        self.assertEqual(first.tipo_servico.nome, "Manutenção de Esquadrias")
        self.assertEqual(first.servico.nome, "Porta de Madeira")
        self.assertEqual(first.gut_score, 48)
        self.assertEqual(first.gut_nivel, "REGULAR")
        self.assertEqual(str(first.predio.latitude), "-7.123456")
        self.assertEqual(str(first.predio.longitude), "-34.654321")
        self.assertEqual(first.contato_direto_url, "(83) 99999-9999")
        self.assertEqual(first.dias_desde_abertura, (date.today() - first.data_cadastro).days)

        second = Requisicao.objects.get(codigo="200/2026")
        self.assertEqual(second.situacao_requisicao, "Inativa")
        self.assertEqual(second.prioridade_final, "2 - Alta")
        self.assertEqual(second.dias_para_execucao, 5)
        self.assertEqual(second.dias_desde_abertura, (date.today() - second.data_cadastro).days)
        self.assertTrue(
            TaxonomiaServico.objects.filter(
                divisao="Construção Civil",
                tipo_servico="Manutenção de Esquadrias",
                servico="Porta de Madeira",
            ).exists()
        )

    def test_import_summary_contains_counts_used_for_parity_checks(self):
        uploaded = self.build_workbook_file()
        result = WorkbookImporter().import_file(uploaded)

        counts = result.resumo_json["contagens"]
        self.assertEqual(counts["divisao"]["Construção Civil"], 1)
        self.assertEqual(counts["divisao"]["Instalações Hidráulicas e Sanitárias"], 1)
        self.assertEqual(counts["status_sipac"]["04 OS EMITIDA"], 1)
        self.assertEqual(counts["status_sipac"]["06 FINALIZADA"], 1)
        self.assertEqual(counts["prioridade_final"]["1 - Urgente"], 1)
        self.assertEqual(counts["prioridade_final"]["2 - Alta"], 1)

    def test_import_csv_handles_windows1252_encoding_and_semicolon_delimiter(self):
        # Reproduz o 500 de produção: CSV exportado pelo Excel/SIPAC em
        # Windows-1252 (byte 0xC7 = "Ç") e delimitado por ponto-e-vírgula.
        header = "Nº Requisição;Assunto;Divisão;Requisitante"
        linha = "100/2026;MANUTENÇÃO de Construção;Construção Civil;MARIA TESTE"
        conteudo = (header + "\n" + linha + "\n").encode("cp1252")
        self.assertIn(b"\xc7", conteudo)

        uploaded = SimpleUploadedFile("importacao.csv", conteudo, content_type="text/csv")
        result = WorkbookImporter().import_file(uploaded)

        self.assertEqual(result.resumo_json["total_processado"], 1)
        requisicao = Requisicao.objects.get(codigo="100/2026")
        self.assertEqual(requisicao.assunto, "MANUTENÇÃO de Construção")
        self.assertEqual(requisicao.divisao.nome, "Construção Civil")

    def test_import_workbook_loads_status_options_from_legendas_1_column_w(self):
        uploaded = self.build_workbook_file()
        WorkbookImporter().import_file(uploaded)

        self.assertTrue(StatusRequisicao.objects.filter(codigo="04 OS EMITIDA").exists())
        self.assertTrue(StatusRequisicao.objects.filter(codigo="06 FINALIZADA").exists())
        self.assertEqual(StatusRequisicao.objects.get(codigo="04 OS EMITIDA").numero, "04")
        self.assertEqual(StatusRequisicao.objects.get(codigo="04 OS EMITIDA").nome, "OS emitida")


class CadastroLoteImporterTests(TestCase):
    def build_lote_file(self, linhas, *, columns=None, name="cadastro-lote.xlsx"):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Requisições"
        sheet.append(columns if columns is not None else list(CADASTRO_LOTE_COLUNAS.values()))
        for linha in linhas:
            sheet.append(linha)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            name,
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def linha_valida(self):
        return [
            "123/2026",
            "Troca de lâmpada queimada",
            "19/05/2026",
            "Elétrica",
            "Iluminação",
            "Troca de lâmpada",
            "01 ABERTA",
            "Bloco CT",
            "João Silva",
            "Coordenação de Exemplo",
        ]

    def test_modelo_xlsx_contem_cabecalho_esperado(self):
        workbook = WorkbookImporter().build_modelo_cadastro_lote()
        sheet = workbook["Requisições"]
        header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        self.assertEqual(header, list(CADASTRO_LOTE_COLUNAS.values()))
        self.assertIn("Instruções", workbook.sheetnames)

    def test_importa_validas_e_relata_invalidas(self):
        linha_sem_assunto = self.linha_valida()
        linha_sem_assunto[0] = "200/2026"
        linha_sem_assunto[1] = ""
        linha_data_ruim = self.linha_valida()
        linha_data_ruim[0] = "300/2026"
        linha_data_ruim[2] = "32/13/2026"
        linha_codigo_ruim = self.linha_valida()
        linha_codigo_ruim[0] = "sem-numero"

        uploaded = self.build_lote_file(
            [self.linha_valida(), linha_sem_assunto, linha_data_ruim, linha_codigo_ruim]
        )
        importacao = WorkbookImporter().import_cadastro_lote(uploaded)
        resumo = importacao.resumo_json

        self.assertEqual(resumo["total_processado"], 1)
        self.assertEqual(resumo["linhas_com_erro"], 3)
        self.assertEqual(Requisicao.objects.count(), 1)
        self.assertTrue(Requisicao.objects.filter(codigo="123/2026").exists())

        linhas_erro = {item["linha"]: item["problemas"] for item in resumo["erros_linhas"]}
        self.assertIn(3, linhas_erro)  # linha do assunto vazio (linha 1 = cabeçalho)
        self.assertTrue(any("Assunto" in p for p in linhas_erro[3]))
        self.assertTrue(any("Data de Cadastro" in p for p in linhas_erro[4]))
        self.assertTrue(any("Nº Requisição" in p for p in linhas_erro[5]))

    def test_modelo_inclui_colunas_opcionais(self):
        workbook = WorkbookImporter().build_modelo_cadastro_lote()
        header = [cell.value for cell in next(workbook["Requisições"].iter_rows(max_row=1))]
        for opcional in ("Orçamento", "Data de Execução", "Local do Serviço", "Contato",
                          "Link do Atendimento", "Link SIPAC"):
            self.assertIn(opcional, header)

    def test_colunas_opcionais_em_branco_nao_bloqueiam_importacao(self):
        # linha_valida() só preenche as 10 obrigatórias; o cabeçalho tem 16 colunas.
        uploaded = self.build_lote_file([self.linha_valida()])
        importacao = WorkbookImporter().import_cadastro_lote(uploaded)
        self.assertEqual(importacao.resumo_json["total_processado"], 1)
        self.assertEqual(importacao.resumo_json["linhas_com_erro"], 0)

    def test_colunas_opcionais_preenchidas_sao_importadas(self):
        linha = self.linha_valida() + [
            "2.300,00",
            "26/05/2026",
            "Sala 10",
            "(83) 98888-7777",
            "https://exemplo.org/atend/123",
            "https://sipac.ufpb.br/req/123",
        ]
        uploaded = self.build_lote_file([linha])
        WorkbookImporter().import_cadastro_lote(uploaded)

        requisicao = Requisicao.objects.get(codigo="123/2026")
        self.assertEqual(requisicao.local_servico, "Sala 10")
        self.assertEqual(requisicao.data_execucao, date(2026, 5, 26))
        self.assertEqual(requisicao.link_sipac, "https://sipac.ufpb.br/req/123")
        self.assertTrue(requisicao.orcamento)

    def test_modelo_sem_colunas_opcionais_continua_valido(self):
        uploaded = self.build_lote_file(
            [self.linha_valida()],
            columns=list(CADASTRO_LOTE_COLUNAS_OBRIGATORIAS.values()),
        )
        importacao = WorkbookImporter().import_cadastro_lote(uploaded)
        self.assertEqual(importacao.resumo_json["total_processado"], 1)

    def test_arquivo_csv_e_rejeitado_com_mensagem_amigavel(self):
        uploaded = SimpleUploadedFile(
            "dados.csv", b"qualquer;coisa\n", content_type="text/csv"
        )
        with self.assertRaises(ImportErrorPlanilha) as ctx:
            WorkbookImporter().import_cadastro_lote(uploaded)
        self.assertIn("XLSX", str(ctx.exception))

    def test_colunas_faltando_geram_erro_orientando_correcao(self):
        uploaded = self.build_lote_file(
            [["123/2026", "Assunto"]], columns=["Nº Requisição", "Assunto"]
        )
        with self.assertRaises(ImportErrorPlanilha) as ctx:
            WorkbookImporter().import_cadastro_lote(uploaded)
        mensagem = str(ctx.exception)
        self.assertIn("faltando as colunas", mensagem)
        self.assertIn("Divisão", mensagem)

    def test_planilha_so_com_cabecalho_nao_quebra(self):
        uploaded = self.build_lote_file([])
        importacao = WorkbookImporter().import_cadastro_lote(uploaded)
        self.assertEqual(importacao.resumo_json["total_processado"], 0)
        self.assertEqual(importacao.resumo_json["linhas_com_erro"], 0)
