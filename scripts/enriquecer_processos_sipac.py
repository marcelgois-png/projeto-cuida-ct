from __future__ import annotations

import argparse
import html
import re
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Codex"


class TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._table_stack: list[list[list[str]]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "table":
            self._table_stack.append([])
        elif tag == "tr" and self._table_stack:
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            self._cell = []

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self._cell is not None and self._row is not None:
            text = clean_text(" ".join(self._cell))
            self._row.append(text)
            self._cell = None
        elif tag == "tr" and self._row is not None and self._table_stack:
            if any(self._row):
                self._table_stack[-1].append(self._row)
            self._row = None
        elif tag == "table" and self._table_stack:
            self.tables.append(self._table_stack.pop())


@dataclass
class SipacData:
    unidade_origem: str
    interessados: list[tuple[str, str, str]]
    erro: str = ""


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(value or "")).strip()


def fetch(url: str, timeout: int = 30) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read().decode("utf-8", "replace")


def extract_unidade_origem(document: str) -> str:
    text = clean_text(re.sub(r"<[^>]+>", " ", document))
    match = re.search(r"Unidade de Origem:\s*(.*?)\s*(?:Status:|Data de Cadastro:|Observação:)", text)
    return clean_text(match.group(1)) if match else ""


def extract_interessados(document: str) -> list[tuple[str, str, str]]:
    parser = TableParser()
    parser.feed(document)

    for table in parser.tables:
        if not table:
            continue
        header = [clean_text(cell).lower() for cell in table[0]]
        if header[:3] == ["tipo", "identificador", "nome"]:
            rows: list[tuple[str, str, str]] = []
            for row in table[1:]:
                padded = (row + ["", "", ""])[:3]
                tipo, identificador, nome = (clean_text(value) for value in padded)
                if tipo or identificador or nome:
                    rows.append((tipo, identificador, nome))
            return rows
    return []


def extract_sipac(url: str) -> SipacData:
    try:
        document = fetch(url)
        return SipacData(
            unidade_origem=extract_unidade_origem(document),
            interessados=extract_interessados(document),
        )
    except (urllib.error.URLError, TimeoutError, OSError) as exc:
        return SipacData(unidade_origem="", interessados=[], erro=str(exc))


def header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None and str(cell.value).strip()
    }


def find_append_column(ws, header: str) -> int:
    headers = header_map(ws)
    if header in headers:
        return headers[header]
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=header)
    return col


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, max_width: int = 80) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for row_idx in range(1, min(ws.max_row, 300) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--pause", type=float, default=0.2)
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(input_path)
    if "PROCESSOS" not in wb.sheetnames:
        raise SystemExit("A aba PROCESSOS não foi encontrada.")

    ws = wb["PROCESSOS"]
    headers = header_map(ws)
    processo_col = headers.get("numero_processo")
    link_col = headers.get("link_sipac")
    if not processo_col or not link_col:
        raise SystemExit("As colunas numero_processo e link_sipac são obrigatórias.")

    unidade_col = find_append_column(ws, "unidade_origem")

    cache: dict[str, SipacData] = {}
    interessados_seen: set[tuple[str, str, str, str]] = set()
    interessados_rows: list[tuple[str, str, str, str]] = []
    erros: list[tuple[int, str, str]] = []

    total = ws.max_row - 1
    for row_idx in range(2, ws.max_row + 1):
        numero = clean_text(str(ws.cell(row=row_idx, column=processo_col).value or ""))
        url = clean_text(str(ws.cell(row=row_idx, column=link_col).value or ""))
        if not url:
            continue

        if url not in cache:
            cache[url] = extract_sipac(url)
            time.sleep(args.pause)

        data = cache[url]
        ws.cell(row=row_idx, column=unidade_col, value=data.unidade_origem)

        if data.erro:
            erros.append((row_idx, numero, data.erro))

        for tipo, identificador, nome in data.interessados:
            key = (numero, tipo, identificador, nome)
            if key not in interessados_seen:
                interessados_seen.add(key)
                interessados_rows.append(key)

        if row_idx % 25 == 0 or row_idx == ws.max_row:
            print(f"Processadas {row_idx - 1}/{total} linhas")

    sheet_name = "INTERESSADOS_PROCESSO"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    out = wb.create_sheet(sheet_name)
    out.append(["numero_processo", "tipo", "identificador", "nome"])
    for row in interessados_rows:
        out.append(row)

    style_header(ws)
    style_header(out)
    autosize(ws)
    autosize(out)

    if erros:
        err_sheet = "ERROS_SIPAC"
        if err_sheet in wb.sheetnames:
            del wb[err_sheet]
        err = wb.create_sheet(err_sheet)
        err.append(["linha_PROCESSOS", "numero_processo", "erro"])
        for row in erros:
            err.append(row)
        style_header(err)
        autosize(err)

    wb.save(output_path)
    print(f"Arquivo salvo: {output_path}")
    print(f"Links únicos consultados: {len(cache)}")
    print(f"Interessados coletados: {len(interessados_rows)}")
    print(f"Erros de coleta: {len(erros)}")


if __name__ == "__main__":
    main()
